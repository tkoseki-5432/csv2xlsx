from pathlib import Path
import pandas as pd
import openpyxl as xl
import os
from openpyxl.chart import LineChart, Reference, Series

#global variable
num = 0
threshold_el = 1
threshold_az = 0.3
A2Cnum = list()
sOrgPath = os.getcwd()

#make excel fail
wb1 = xl.Workbook()
sheet1 = wb1.create_sheet(title = "data_el")
sheet2 = wb1.create_sheet(title = "data_Az")
sheet3 = wb1.create_sheet(title = "summry")

#conti result excel open
wb2 = xl.load_workbook("Conti_Results.xlsx")
sheet_name = wb2.get_sheet_names()
sheet = wb2.get_sheet_by_name(str(sheet_name))

#get csv file list
path_obj = Path()

csvfile = list(path_obj.glob("*.csv"))

#copy az and el data form csvfile to ExcelFile
for listnum in  range(len(csvfile)):
    
    #get A2C No.
    name = str(csvfile[listnum])
    #get Csvdata
    CsvData = pd.read_csv(str(csvfile[listnum]),encoding="SHIFT-JIS", skiprows=2)
    
    #同一のA2C番号があった場合は、その行に書き込む。ない場合は、新しい行に書き込む。
    if A2Cnum.count(name[15:23])==0:
        #write A2Cnum to xlsx
        A2Cnum.append(name[15:23])
        num = len(A2Cnum)
        sheet1.cell(row=1, column=num+1).value = A2Cnum[num-1] 
        sheet2.cell(row=1, column=num+1).value = A2Cnum[num-1] 
     
    else:
        num=A2Cnum.index(name[15:23])+1

    for p in range(len(CsvData['Error-P'])):
        
        sheet1.cell(row=CsvData.iat[p,1]+1, column =num+1).value = CsvData['Error-P'][p]
        sheet2.cell(row=CsvData.iat[p,1]+1, column =num+1).value = CsvData['Error-Y'][p]
        
        #Pitch,Yaw,Rollを左端に追加
        Pitch=int(CsvData['RAD-P'][p])
        Yaw=int(CsvData['RAD-Y'][p])
        Roll=int(CsvData['RAD-R'][p])
        sheet1.cell(row=CsvData.iat[p,1]+1, column =1).value = 'P{0}Y{1}R{2}'.format(Pitch,Yaw,Roll)
        sheet2.cell(row=CsvData.iat[p,1]+1, column =1).value = 'P{0}Y{1}R{2}'.format(Pitch,Yaw,Roll)

#write threshold to xlsx
for i in range(1,3):
    sheet1.cell(row=1, column=sheet1.max_column+1).value = "thresh"
    sheet2.cell(row=1, column=sheet2.max_column+1).value = "thresh" 
        
    for j in range(2,sheet1.max_row):
        sheet1.cell(row=j, column=sheet1.max_column).value = threshold_el*(-1)**i
        sheet2.cell(row=j, column=sheet1.max_column).value = threshold_az*(-1)**i

#----------------make chart---------------#
#make chart at sheet1
values = Reference(sheet1,min_col = 2,min_row=1,max_col=sheet1.max_column, max_row=sheet1.max_row)
cats  = Reference(sheet1,min_col=1,min_row=2,max_row=sheet1.max_row)
chart=LineChart()

chart.add_data(values,titles_from_data=True)
chart.set_categories(cats)

##setting
s1=chart.series[-1]
s1.graphicalProperties.line.solidFill = "E00E00"
s2=chart.series[-2]
s2.graphicalProperties.line.solidFill = "E00E00"

sheet1.add_chart(chart,'A1')

#make chart at sheet2
values = Reference(sheet2,min_col = 2,min_row=1,max_col=sheet1.max_column, max_row=sheet2.max_row)
cats  = Reference(sheet2,min_col=1,min_row=2,max_row=sheet2.max_row)
chart=LineChart()

chart.add_data(values,titles_from_data=True)
chart.set_categories(cats)

##setting
s1=chart.series[-1]
s1.graphicalProperties.line.solidFill = "E00E00"
s2=chart.series[-2]
s2.graphicalProperties.line.solidFill = "E00E00"
sheet2.add_chart(chart,'A1')

wb1.save("file.xlsx")
